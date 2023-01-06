import csv
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00


def create_table(work_sheet, tags, collection, index_table, is_percent):
    side = Side(style='thin')
    border = Border(left=side, right=side, top=side, bottom=side)
    for index in range(len(tags)):
        work_sheet.cell(row=1, column=index + index_table).value = tags[index]
        work_sheet.cell(row=1, column=index + index_table).font = Font(bold=True)
        work_sheet.cell(row=1, column=index + index_table).border = border
    for element in collection:
        for index in range(len(element)):
            work_sheet.cell(row=index + 2, column=collection.index(element) + index_table).value = element[index]
            work_sheet.cell(row=index + 2, column=collection.index(element) + index_table).border = border
            if is_percent:
                work_sheet.cell(row=index + 2,
                                column=collection.index(element) + 4).number_format = FORMAT_PERCENTAGE_00
    return fix_width_table(work_sheet, is_percent)


def fix_width_table(work_sheet, is_second_table):
    dims = {}
    for row in work_sheet.rows:
        for cell in row:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    if is_second_table:
        dims['C'] = 0
    for key, width in dims.items():
        work_sheet.column_dimensions[key].width = width + 2
    return work_sheet


class DataSet:
    def __init__(self):
        self.file_name = input("Введите название файла: ")
        self.profession_name = input("Введите название профессии: ")
        currency_name = ["AZN", "BYR", "EUR", "GEL", "KGS", "KZT", "RUR", "UAH", "USD", "UZS"]
        currency_value = [35.68, 23.91, 59.90, 21.74, 0.76, 0.13, 1, 1.64, 60.66, 0.0055]
        self.currency_to_rub = dict(zip(currency_name, currency_value))
        self.professions = []
        self.year_collection = []
        self.city_collection = []

    def get_correct_data(self):
        with open(self.file_name, encoding='utf-8-sig') as file:
            for row in csv.DictReader(file):
                correct_row = True
                for field in row:
                    if row[field] is None or len(row[field]) == 0:
                        correct_row = False
                        continue
                if correct_row:
                    self.professions.append(Profession(*[row[field] for field in row]))

    def get_info(self):
        salary_dict = dict()
        count_dict = dict()
        for profession in self.professions:
            profession.salary_from = int(str(profession.salary_from).replace('.0', ''))
            profession.salary_to = int(str(profession.salary_to).replace('.0', ''))
            if profession.get_year() not in salary_dict.keys():
                salary_dict[profession.get_year()] = []
            salary_dict[profession.get_year()].append((profession.salary_from + profession.salary_to) // 2
                                                      * self.currency_to_rub[profession.salary_currency])
        for key, value in salary_dict.items():
            salary_dict[key] = int(str((sum(value) // len(value))).replace('.0', ''))
            count_dict[key] = len(value)
        print("Динамика уровня зарплат по годам:", salary_dict)
        print('Динамика количества вакансий по годам:', count_dict)
        self.year_collection.append(salary_dict)
        self.year_collection.append(count_dict)

    def get_info_professions(self):
        salary_dict = dict()
        count_dict = dict()
        for profession in self.professions:
            if profession.get_year() not in salary_dict.keys():
                salary_dict[profession.get_year()] = []
            if self.profession_name in profession.name:
                profession.salary_from = int(str(profession.salary_from).replace('.0', ''))
                profession.salary_to = int(str(profession.salary_to).replace('.0', ''))
                salary_dict[profession.get_year()].append((profession.salary_from + profession.salary_to) // 2
                                                          * self.currency_to_rub[profession.salary_currency])
        for key, value in salary_dict.items():
            if len(value) != 0:
                salary_dict[key] = int(str((sum(value) // len(value))).replace('.0', ''))
                count_dict[key] = len(value)
            else:
                salary_dict[key] = 0
                count_dict[key] = 0

        print("Динамика уровня зарплат по годам для выбранной профессии:", salary_dict)
        print('Динамика количества вакансий по годам для выбранной профессии:', count_dict)
        self.year_collection.append(salary_dict)
        self.year_collection.append(count_dict)

    def get_info_cities(self):
        city_dict = dict()
        ans_dict = dict()
        count_vacancy = 0
        for profession in self.professions:
            profession.salary_from = int(str(profession.salary_from).replace('.0', ''))
            profession.salary_to = int(str(profession.salary_to).replace('.0', ''))
            if profession.area_name not in city_dict.keys():
                city_dict[profession.area_name] = []
            city_dict[profession.area_name].append((profession.salary_from + profession.salary_to) // 2
                                                   * self.currency_to_rub[profession.salary_currency])
            count_vacancy += 1
        percent_dict = dict()
        for key, value in city_dict.items():
            if len(value) >= count_vacancy // 100:
                ans_dict[key] = int(str((sum(value) // len(value))).replace('.0', ''))
                percent_dict[key] = round(len(value) / count_vacancy, 4)
        ans_dict = dict(sorted(ans_dict.items(), key=lambda item: item[1], reverse=True)[:10])
        percent_dict = dict(sorted(percent_dict.items(), key=lambda item: item[1], reverse=True)[:10])
        print('Уровень зарплат по городам (в порядке убывания):', ans_dict)
        print('Доля вакансий по городам (в порядке убывания):', percent_dict)
        self.city_collection.append(ans_dict)
        self.city_collection.append(percent_dict)


class Profession:
    def __init__(self, name, salary_from, salary_to, salary_currency, area_name, published_at):
        self.name = name
        self.salary_from = int(salary_from.replace('.0', ''))
        self.salary_to = int(salary_to.replace('.0', ''))
        self.salary_currency = salary_currency
        self.area_name = area_name
        self.published_at = published_at

    def get_year(self):
        return int(self.published_at[:4])


class Report:
    def __init__(self):
        self.book = openpyxl.Workbook()
        self.sheet = self.book.active
        self.sheet.title = "Статистика по годам"
        self.second_sheet = self.book.create_sheet('Статистика по городам')

    def generate_excel(self, year_collection, city_collection):
        collection_level = [list(city_collection[0].keys())] + [list(city_collection[0].values())]
        collection_fraction = [list(city_collection[1].keys())] + [list(city_collection[1].values())]
        collection_years = [list(year_collection[0].keys())]
        [collection_years.append(list(data.values())) for data in year_collection]
        collection_years = [collection_years[0]] + [collection_years[1]] + [collection_years[3]] + \
                           [collection_years[2]] + [collection_years[4]]
        names = ["Год", "Средняя зарплата", f"Средняя зарплата - {data_professions.profession_name}",
                 "Количество вакансий", f"Количество вакансий - {data_professions.profession_name}"]
        self.sheet = create_table(self.sheet, names, collection_years, 1, False)
        self.second_sheet = create_table(self.second_sheet, ["Город", "Уровень зарплат"], collection_level, 1, False)
        self.second_sheet = create_table(self.second_sheet, ["Город", "Доля вакансий"], collection_fraction, 4, True)
        self.book.save("report.xlsx")
        self.book.close()


data_professions = DataSet()
data_professions.get_correct_data()
data_professions.get_info()
data_professions.get_info_professions()
data_professions.get_info_cities()

report = Report()
report.generate_excel(data_professions.year_collection, data_professions.city_collection)
