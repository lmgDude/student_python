import csv
import os
import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from matplotlib import pyplot as plt
import pdfkit
from jinja2 import Environment, FileSystemLoader


def create_table(work_sheet, tags, collection, index_table, is_percent):
    side = Side(style='thin')
    border = Border(left=side, right=side, top=side, bottom=side)
    for index in range(len(tags)):
        work_sheet.cell(row=1, column=index + index_table).value = tags[index]
        work_sheet.cell(row=1, column=index + index_table).font = Font(bold=True)
        work_sheet.cell(row=1, column=index + index_table).border = border
    for element in collection:
        for index in range(len(element)):
            work_sheet.cell(row=index + 2, column=collection.index(element) + index_table).value = element[index]
            work_sheet.cell(row=index + 2, column=collection.index(element) + index_table).border = border
            if is_percent:
                work_sheet.cell(row=index + 2,
                                column=collection.index(element) + 4).number_format = FORMAT_PERCENTAGE_00
    return fix_width_table(work_sheet, is_percent)


def fix_width_table(work_sheet, is_second_table):
    dims = {}
    for row in work_sheet.rows:
        for cell in row:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    if is_second_table:
        dims['C'] = 0
    for key, width in dims.items():
        work_sheet.column_dimensions[key].width = width + 2
    return work_sheet


class DataSet:
    def __init__(self):
        self.file_name = input("Введите название файла: ")
        self.profession_name = input("Введите название профессии: ")
        currency_name = ["AZN", "BYR", "EUR", "GEL", "KGS", "KZT", "RUR", "UAH", "USD", "UZS"]
        currency_value = [35.68, 23.91, 59.90, 21.74, 0.76, 0.13, 1, 1.64, 60.66, 0.0055]
        self.currency_to_rub = dict(zip(currency_name, currency_value))
        self.professions = []
        self.year_collection = []
        self.city_collection = []

    def get_correct_data(self):
        with open(self.file_name, encoding='utf-8-sig') as file:
            for row in csv.DictReader(file):
                correct_row = True
                for field in row:
                    if row[field] is None or len(row[field]) == 0:
                        correct_row = False
                        continue
                if correct_row:
                    self.professions.append(Profession(*[row[field] for field in row]))

    def get_info(self, is_profession):
        # Если нужно получить информацию по годам для выбранной профессии => is_profession нужно установить на True

        salary_dict = dict()
        count_dict = dict()
        for profession in self.professions:
            profession.salary_from = int(profession.salary_from)
            profession.salary_to = int(profession.salary_to)

            if profession.get_year() not in salary_dict.keys():
                salary_dict[profession.get_year()] = []

            if is_profession:
                if self.profession_name in profession.name:
                    salary_dict[profession.get_year()].append((profession.salary_from + profession.salary_to) // 2
                                                              * self.currency_to_rub[profession.salary_currency])
            else:
                salary_dict[profession.get_year()].append(int((profession.salary_from + profession.salary_to) // 2
                                                              * self.currency_to_rub[profession.salary_currency]))

        for key, value in salary_dict.items():
            salary_dict[key] = int(sum(value) // len(value)) if len(value) != 0 else 0
            count_dict[key] = len(value) if len(value) != 0 else 0

        if is_profession:
            print("Динамика уровня зарплат по годам для выбранной профессии:", salary_dict)
            print('Динамика количества вакансий по годам для выбранной профессии:', count_dict)
        else:
            print("Динамика уровня зарплат по годам:", salary_dict)
            print('Динамика количества вакансий по годам:', count_dict)

        self.year_collection.append(salary_dict)
        self.year_collection.append(count_dict)

    def get_info_cities(self):
        city_dict = dict()
        ans_dict = dict()
        count_vacancy = 0

        for profession in self.professions:
            profession.salary_from = int(profession.salary_from)
            profession.salary_to = int(profession.salary_to)

            if profession.area_name not in city_dict.keys():
                city_dict[profession.area_name] = []
            city_dict[profession.area_name].append(int((profession.salary_from + profession.salary_to) // 2
                                                       * self.currency_to_rub[profession.salary_currency]))
            count_vacancy += 1
        percent_dict = dict()

        for key, value in city_dict.items():
            if len(value) >= count_vacancy // 100:
                ans_dict[key] = int(sum(value) // len(value))
                percent_dict[key] = round(len(value) / count_vacancy, 4)

        ans_dict = dict(sorted(ans_dict.items(), key=lambda item: item[1], reverse=True)[:10])
        percent_dict = dict(sorted(percent_dict.items(), key=lambda item: item[1], reverse=True)[:10])

        print('Уровень зарплат по городам (в порядке убывания):', ans_dict)
        print('Доля вакансий по городам (в порядке убывания):', percent_dict)

        self.city_collection.append(ans_dict)
        self.city_collection.append(percent_dict)

    def get_general_info(self):
        self.get_correct_data()
        self.get_info(False)
        self.get_info(True)
        self.get_info_cities()


class Profession:
    def __init__(self, name, salary_from, salary_to, salary_currency, area_name, published_at):
        self.name = name
        self.salary_from = int(salary_from.replace('.0', ''))
        self.salary_to = int(salary_to.replace('.0', ''))
        self.salary_currency = salary_currency
        self.area_name = area_name
        self.published_at = published_at

    def get_year(self):
        return int(self.published_at[:4])


class Report:
    def __init__(self, data):
        self.book = openpyxl.Workbook()
        self.sheet = self.book.active
        self.sheet.title = "Статистика по годам"
        self.second_sheet = self.book.create_sheet('Статистика по городам')
        self.data = data
        self.year_collection = data.year_collection
        self.city_collection = data.city_collection

    def generate_excel(self):
        collection_level = [list(self.city_collection[0].keys())] + [list(self.city_collection[0].values())]
        collection_fraction = [list(self.city_collection[1].keys())] + [list(self.city_collection[1].values())]
        collection_years = [list(self.year_collection[0].keys())]
        [collection_years.append(list(data.values())) for data in self.year_collection]
        collection_years = [collection_years[0]] + [collection_years[1]] + [collection_years[3]] + \
                           [collection_years[2]] + [collection_years[4]]
        names = ["Год", "Средняя зарплата", f"Средняя зарплата - {self.data.profession_name}",
                 "Количество вакансий", f"Количество вакансий - {self.data.profession_name}"]
        self.sheet = create_table(self.sheet, names, collection_years, 1, False)
        self.second_sheet = create_table(self.second_sheet, ["Город", "Уровень зарплат"], collection_level, 1, False)
        self.second_sheet = create_table(self.second_sheet, ["Город", "Доля вакансий"], collection_fraction, 4, True)
        self.book.save("template/report.xlsx")
        self.book.close()

    def generate_image(self):
        figure, axis = plt.subplots(2, 2)
        width = 0.4
        profession = self.data.profession_name.lower()
        self.city_collection[1]['Другие'] = 1 - sum(list(self.city_collection[1].values()))

        salary_and_vacancy_year_x = list(self.year_collection[0].keys())
        salary_year_y = list(self.year_collection[0].values())
        salary_year_profession_y = list(self.year_collection[2].values())

        vacancy_year_y = list(self.year_collection[1].values())
        vacancy_year_profession_y = list(self.year_collection[3].values())
        x = np.arange(len(salary_and_vacancy_year_x))

        salary_city_y = [x.replace(' ', '\n').replace('-', '-\n') for x in list(self.city_collection[0].keys())]
        salary_city_x = list(self.city_collection[0].values())

        vacancy_city_values = list(self.city_collection[1].values())
        vacancy_city_labels = list(self.city_collection[1].keys())

        axis[0, 0].bar(x - width / 2, salary_year_y, width)
        axis[0, 0].bar(x + width / 2, salary_year_profession_y, width)
        axis[0, 0].set_xticks(x, salary_and_vacancy_year_x)
        axis[0, 0].tick_params(axis='x', labelsize=8, rotation=90)
        axis[0, 0].tick_params(axis='y', labelsize=8)
        axis[0, 0].grid(axis='y')
        axis[0, 0].set_title("Уровень зарплат по годам")
        axis[0, 0].legend(['средняя з/п', f'з/п {profession}'], fontsize=8)

        axis[0, 1].bar(x - width / 2, vacancy_year_y, width)
        axis[0, 1].bar(x + width / 2, vacancy_year_profession_y, width)
        axis[0, 1].set_xticks(x, salary_and_vacancy_year_x)
        axis[0, 1].tick_params(axis='x', labelsize=8, rotation=90)
        axis[0, 1].tick_params(axis='y', labelsize=8)
        axis[0, 1].grid(axis='y')
        axis[0, 1].set_title("Количество вакансий по годам", fontsize=11)
        axis[0, 1].legend(['Количество вакансий', f'Количество вакансий\n{profession}'], fontsize=8)

        axis[1, 0].barh(salary_city_y, salary_city_x)
        axis[1, 0].invert_yaxis()
        axis[1, 0].tick_params(axis='x', labelsize=8)
        axis[1, 0].tick_params(axis='y', labelsize=6)
        axis[1, 0].grid(axis='x')
        axis[1, 0].set_title("Уровень зарплат по городам")
        axis[1, 0].plot()

        axis[1, 1].pie(vacancy_city_values, labels=vacancy_city_labels, textprops={'fontsize': 6})
        axis[1, 1].set_title("Доля вакансий по городам")
        axis[1, 1].plot()

        plt.tight_layout()
        plt.savefig('template/graph.png')

    def generate_pdf(self):
        path = r'Путь к html'
        path_second = r'Путь к шаблону 2'
        path_third = r'Путь к шаблону 3'
        excel_path = r'Путь к exel'

        first_table = pd.read_excel(excel_path)
        second_table = pd.read_excel(excel_path,
                                     sheet_name='Статистика по городам').iloc[:, [True, True, False, False, False]]
        third_table = pd.read_excel(excel_path,
                                    sheet_name='Статистика по городам').iloc[:, [False, False, False, True, True]]
        third_table.columns = ['Город', 'Доля вакансий']
        third_table['Доля вакансий'] = third_table['Доля вакансий'] * 100

        [table.to_html(element, index=False) for table, element in [[first_table, path], [second_table, path_second],
                                                                    [third_table, path_third]]]
        template = Environment(loader=FileSystemLoader('.')).get_template("template/result.html")
        pdf_template = template.render({'name': self.data.profession_name,
                                        'first_table': open(path, 'r', encoding='utf-8-sig').read(),
                                        'second_table': open(path_second, 'r', encoding='utf-8-sig').read(),
                                        'third_table': open(path_third, 'r', encoding='utf-8-sig').read()})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'template/report.pdf', configuration=config,
                           options={"enable-local-file-access": ""})
        [os.remove(element) for element in [path, path_second, path_third]]

    def execute_report(self):
        self.generate_excel()
        self.generate_image()
        self.generate_pdf()


def execute():
    data_professions = DataSet()
    data_professions.get_general_info()


execute()